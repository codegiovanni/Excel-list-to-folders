import os
import openpyxl

RESULT_LOCATION = r"enter full path\result"
DATA = r"enter full path\data\excel file name.xlsx"

workbook = openpyxl.load_workbook(DATA)
sheet = workbook['Sheet1']

# Use this part when folder name is created just from one column
# --------------------------------------------------------------
# column_values = [cell.value for col in sheet.iter_cols(
#     min_row=2, max_row=None, min_col=2, max_col=2) for cell in col]
# column_values = list(dict.fromkeys(column_values))  # removes duplicates
#
# for value in column_values:
#     print("Creating folder: ", value)

# for value in column_values:
#     folderName = value
#     baseDir = RESULT_LOCATION
#     os.makedirs(os.path.join(baseDir, folderName))
#     print("Created folder: ", folderName)
# --------------------------------------------------------------


# Use this part when folder name is created from several columns
# --------------------------------------------------------------
column_values = [(cell.value for col in sheet.iter_cols(
    min_row=2, max_row=None, min_col=1, max_col=1) for cell in col), (cell.value for col in sheet.iter_cols(
    min_row=2, max_row=None, min_col=2, max_col=2) for cell in col), (cell.value for col in sheet.iter_cols(
    min_row=2, max_row=None, min_col=3, max_col=3) for cell in col)]

column_1 = column_values[0]
column_2 = column_values[1]
column_3 = column_values[2]

result = ("{}_{}_{}".format(x, y, z) for x, y, z in zip(column_1, column_2, column_3))
result = list(dict.fromkeys(result))  # removes duplicates

# for value in result:
#     print("Creating folder: ", value)

for value in result:
    folderName = value
    baseDir = RESULT_LOCATION
    os.makedirs(os.path.join(baseDir, folderName))
    print("Created folder: ", folderName)
# --------------------------------------------------------------
